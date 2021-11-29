from time import time
import tkinter as tk
import openpyxl
window = tk.Tk()

window.title("Time study")

window.rowconfigure(0, minsize=400, weight=1)
window.columnconfigure(1, minsize=400, weight=1)

txt_label = tk.Label(window)
fr_buttons = tk.Frame(window)

data = []


def cmd_occupied():
    txt_label.config(text="occupied")
    data.append({"timestamp": round(time() * 1000), "occupied": True})


def cmd_free():
    txt_label.config(text="free")
    data.append({"timestamp": round(time() * 1000), "occupied": False})


btn_occupied = tk.Button(fr_buttons, text="Occupied", command=cmd_occupied)
btn_free = tk.Button(fr_buttons, text="Free", command=cmd_free)

btn_free.grid(row=0, column=0, sticky="ew", padx=5, pady=5)
btn_occupied.grid(row=1, column=0, sticky="ew", padx=5)
fr_buttons.grid(row=0, column=0, sticky="ns")
txt_label.grid(row=0, column=1, sticky="nsew")

window.mainloop()

for i in data:
    print(i)

wb = openpyxl.Workbook()
sheet = wb.active
for i, val in enumerate(data):
    sheet.cell(column=1, row=i+1, value=val.get("timestamp"))
    sheet.cell(column=2, row=i+1, value=val.get("occupied"))
wb.save(f"manual_overlay_{int(time()*1000)}.xlsx")
