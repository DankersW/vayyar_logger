from pathlib import Path
import openpyxl


class ManualOverlay:
    def __init__(self, filename: Path):
        self.filename = filename
        self._data = self._parse_data()

    def _parse_data(self) -> list:
        xlsx = openpyxl.load_workbook(self.filename).active
        data = []
        start_t = xlsx.cell(row=1, column=1, value=None).value
        data.append({"timestamp": start_t-1, "room_occupied": False})
        for row in range(xlsx.max_row):
            timestamp = xlsx.cell(row=row+1, column=1, value=None).value
            room_status = xlsx.cell(row=row+1, column=2, value=None).value

            if isinstance(room_status, bool):
                status = room_status
            else:
                status = True if room_status.startswith("=TRUE()") else False

            prev_item = data[-1].get("room_occupied")
            if prev_item != status:
                data.append({"timestamp": timestamp-1, "room_occupied": prev_item})
            data.append({"timestamp": timestamp, "room_occupied": status})

        for i in data:
            print(i)
        #exit(1)
        return data

    @staticmethod
    def _create_entry(ts: int, status: any) -> dict:
        entry = {"timestamp": ts}
        if isinstance(status, bool):
            entry["room_occupied"] = status
        else:
            if status.startswith("=TRUE()"):
                entry["room_occupied"] = True
            else:
                entry["room_occupied"] = False
        return entry

    def get_overlay_data(self) -> dict:
        data = {"data": self._data, "min": self._data[-1].get("timestamp"), "max": self._data[0].get("timestamp"),
                "label": "manual", "color": "blue"}
        return data
