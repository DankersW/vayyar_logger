from typing import Tuple
from datetime import datetime
from pathlib import Path
from dateutil import parser
from dataclasses import dataclass
from time import time

from yaml import safe_load
from matplotlib import pyplot
import openpyxl

from db_driver import DbDriver


class WeldcloudOverlay:
    row_first_data = 7
    col_event_created = 2
    col_arc_time = 3

    def __init__(self, filename: Path):
        self.filename: Path = filename
        self._data = self._parse_data()

    def _parse_data(self) -> list:
        xlsx = openpyxl.load_workbook(self.filename).active
        data = []
        for row in range(self.row_first_data, xlsx.max_row):
            created_time = xlsx.cell(row=row, column=self.col_event_created, value=None).value
            start_timestamp = self._time_to_timestamp(timestamp=created_time)
            arc_time = xlsx.cell(row=row, column=self.col_arc_time, value=None).value
            duration = self._arc_time_to_millisec(arc_time)
            data_enty = self._construct_data_entry(start_timestamp=start_timestamp, duration=duration)
            data.extend(data_enty)
        return data

    def get_overlay_data(self) -> dict:
        data = {"data": self._data, "min": self._data[0].get("timestamp"), "max": self._data[-1].get("timestamp"),
                "label": "weldcloud", "color": "green"}
        return data

    @staticmethod
    def _arc_time_to_millisec(arc_time: str) -> int:
        time_values = arc_time.split(sep=' ')
        duration = 0
        if any("m" in time_value for time_value in time_values):
            duration += int(time_values[0][:-1]) * 60000
        if any("s" in time_value for time_value in time_values):
            duration += int(time_values[-1][:-1]) * 1000
        return duration

    @staticmethod
    def _time_to_timestamp(timestamp: str) -> int:
        if timestamp.isdigit():
            return int(timestamp)
        else:
            return int(parser.parse(timestamp).timestamp()) * 1000

    @staticmethod
    def _construct_data_entry(start_timestamp: int, duration: int) -> list:
        data = []
        if duration != 0:
            data = [{"timestamp": start_timestamp - 1, "welding": 0},
                    {"timestamp": start_timestamp, "welding": 1},
                    {"timestamp": start_timestamp + duration, "welding": 1},
                    {"timestamp": start_timestamp + duration + 1, "welding": 0}]
        return data


class VayyarOverlay:
    def __init__(self, min_ts: int, max_ts: int):
        self.min = min_ts
        self.max = max_ts
        self.db_room = DbDriver(keys=self._get_keys(), table='vayyar_home_c2c_room_status')
        self._data = self._get_room_data(youngest_ts=min_ts, oldest_ts=max_ts)

    def get_overlay_data(self) -> dict:
        data = {"data": self._data, "min": self._data[0].get("timestamp"), "max": self._data[-1].get("timestamp"),
                "label": "vayyar", "color": "orange"}
        return data

    @staticmethod
    def _get_keys():
        with open("keys.yml") as config_file:
            return safe_load(config_file)

    def _get_room_data(self, oldest_ts, youngest_ts) -> list:
        cloud_data = self.db_room.query_db(oldest_timestamp=oldest_ts, untill_timestamp=youngest_ts)
        return self._parse_cloud_data(cloud_data=cloud_data)

    @staticmethod
    def _parse_cloud_data(cloud_data: dict) -> list:
        data = []
        for item in cloud_data:
            timestamp = item.get("timestamp").__int__()
            occupied = 1 if item.get("room_occupied") else 0
            entry = {"timestamp": timestamp, "room_occupied": occupied}
            data.append(entry)
        return data


class PostPlotter:
    def __init__(self, config: dataclass):
        self.overlays = []
        if config.weldcloud_overlay:
            wc_overlay = WeldcloudOverlay(filename=config.weldcloud_data)
            self.overlays.append(wc_overlay.get_overlay_data())

        if config.post_plotter:
            min_ts, max_ts = self.get_min_max_timestamp()
            vayyay_overlay = VayyarOverlay(min_ts, max_ts)
            self.overlays.append(vayyay_overlay.get_overlay_data())
            self.plot_data(config.overlay_counter)

    def get_min_max_timestamp(self) -> (int, int):
        if len(self.overlays) == 0:
            return round(time() * 1000), round(time() * 1000) - 86400000
        min_ts = round(time() * 1000)
        max_ts = 0
        for item in self.overlays:
            if item.get("min") < min_ts:
                min_ts = item.get("min")
            if item.get("max") > max_ts:
                max_ts = item.get("max")
        return min_ts, max_ts

    def plot_data(self, overlay_count: int):
        fig = pyplot.figure()
        combined = fig.add_subplot(overlay_count+1, 1, 1)
        for i, overlay in enumerate(self.overlays):
            sub_plot = fig.add_subplot(overlay_count+1, 1, i+2)
            x, y = self._get_plot_data(data=overlay.get("data"))
            sub_plot.step(x, y, label=overlay.get("label"), color=overlay.get("color"))
            combined.step(x, y, label=overlay.get("label"), color=overlay.get("color"))
            sub_plot.legend()
        combined.legend()
        pyplot.show()

    def _get_plot_data(self, data: list) -> Tuple[list, list]:
        x = []
        y = []
        for item in data:
            _x = self._decimal_timestamp_to_dt(item.get("timestamp"))
            if "welding" in item:
                _y = item.get("welding")
            else:
                _y = item.get("room_occupied")
            x.append(_x)
            y.append(_y)
        return x, y

    @staticmethod
    def _decimal_timestamp_to_dt(timestamp: int) -> datetime:
        return datetime.fromtimestamp(timestamp / 1000)
