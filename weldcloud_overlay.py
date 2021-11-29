from pathlib import Path
from dateutil import parser

import openpyxl


class WeldcloudOverlay:
    row_first_data = 7
    col_event_created = 2
    col_arc_time = 3

    def __init__(self, filename: Path):
        self.filename: Path = filename
        self._data = self._parse_data()

    def _parse_data(self) -> list:
        xlsx = openpyxl.load_workbook(self.filename).active
        data = []
        for row in range(self.row_first_data, xlsx.max_row):
            created_time = xlsx.cell(row=row, column=self.col_event_created, value=None).value
            start_timestamp = self._time_to_timestamp(timestamp=created_time)
            arc_time = xlsx.cell(row=row, column=self.col_arc_time, value=None).value
            duration = self._arc_time_to_millisec(arc_time)
            data_enty = self._construct_data_entry(start_timestamp=start_timestamp, duration=duration)
            data.extend(data_enty)
        return data

    def get_overlay_data(self) -> dict:
        data = {"data": self._data, "min": self._data[0].get("timestamp"), "max": self._data[-1].get("timestamp"),
                "label": "weldcloud", "color": "green"}
        return data

    @staticmethod
    def _arc_time_to_millisec(arc_time: str) -> int:
        time_values = arc_time.split(sep=' ')
        duration = 0
        if any("m" in time_value for time_value in time_values):
            duration += int(time_values[0][:-1]) * 60000
        if any("s" in time_value for time_value in time_values):
            duration += int(time_values[-1][:-1]) * 1000
        return duration

    @staticmethod
    def _time_to_timestamp(timestamp: str) -> int:
        if timestamp.isdigit():
            return int(timestamp)
        else:
            return int(parser.parse(timestamp).timestamp()) * 1000

    @staticmethod
    def _construct_data_entry(start_timestamp: int, duration: int) -> list:
        data = []
        if duration != 0:
            data = [{"timestamp": start_timestamp - 1, "welding": 0},
                    {"timestamp": start_timestamp, "welding": 1},
                    {"timestamp": start_timestamp + duration, "welding": 1},
                    {"timestamp": start_timestamp + duration + 1, "welding": 0}]
        return data
